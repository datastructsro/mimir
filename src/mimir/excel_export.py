"""Export replenishment rules and forecast evidence to Excel (.xlsx) files.

Produces two sheets in one workbook:
  1. "Odoo Import" — columns formatted for direct import via Odoo's
     Inventory > Operations > Replenishment > Import button.
     Uses database IDs (/.id) for product, warehouse, location to
     avoid name-matching ambiguity.
  2. "Review" — human-readable version with product names, lead times,
     and proposed min/max values.
     This sheet is for review only — not meant for Odoo import.

Odoo 19 import format reference:
  - `id` column: external ID for create/update idempotency
  - `product_id/.id`: database ID of the product
  - `warehouse_id/.id`: database ID of the warehouse
  - `location_id/.id`: database ID of the stock location (lot_stock_id)
  - `product_min_qty`, `product_max_qty`: float values
  - `trigger`: 'auto' or 'manual'
  - Odoo accepts .xlsx natively via base_import module (openpyxl)

The external ID pattern `mimir.op_{product_id}_{warehouse_id}` ensures
that re-importing the same file updates existing rules instead of duplicating.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TypeVar, cast

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .schemas import RuleRow

logger = logging.getLogger(__name__)

_Row = TypeVar("_Row")


def _active_ws(wb: Workbook) -> Worksheet:
    """Return a guaranteed-non-None active worksheet.

    openpyxl's Workbook.active is typed Optional because a workbook can
    have zero sheets. In our call sites we just constructed the workbook,
    so it has the default sheet — but pyright can't see that invariant.
    This helper asserts it, creating a new sheet as a last resort.
    """
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    return ws

# Colors for inventory flags
FLAG_FILLS = {
    "ok": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),          # green
    "overstock": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),    # red
    "understock": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # yellow
    "at_risk": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),      # yellow
    "no_stock": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),     # red
    "no_demand": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),    # blue-gray
    "no_data": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),      # gray
}

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")


def export_to_excel(
    rules_parquet: str | Path,
    forecast_parquet: str | Path | None = None,
    decisions_parquet: str | Path | None = None,
    output_path: str | Path | None = None,
) -> Path:
    """Export replenishment rules to an Odoo-importable Excel file.

    Args:
        rules_parquet: Path to replenishment_rules.parquet
        forecast_parquet: Deprecated. Kept for backward compatibility.
        output_path: Output .xlsx path (default: same dir as rules_parquet)

    Returns:
        Path to the written .xlsx file
    """
    rules_parquet = Path(rules_parquet)
    if output_path is None:
        output_path = rules_parquet.parent / "replenishment_rules.xlsx"
    else:
        output_path = Path(output_path)

    # Read rules
    rules_table = pq.read_table(str(rules_parquet))
    rules: list[RuleRow] = _table_to_rows(rules_table, RuleRow)

    # Read decisions for decision_id mapping
    decisions_map: dict[tuple[int, int], str] = {}
    if decisions_parquet and Path(decisions_parquet).exists():
        try:
            dec_table = pq.read_table(
                str(decisions_parquet),
                columns=["_odoo_product_id", "_odoo_location_id", "decision_id"],
            )
            for row in dec_table.to_pylist():
                decisions_map[(row["_odoo_product_id"], row["_odoo_location_id"])] = row["decision_id"]
        except Exception:
            pass

    wb = Workbook()

    # Sheet 1: Odoo Import (clean, minimal, importable)
    _write_odoo_import_sheet(wb, rules)

    # Sheet 2: Review (human-readable with all detail)
    _write_review_sheet(wb, rules, decisions_map)

    wb.save(str(output_path))
    logger.info("Wrote Excel export to %s (%d rules)", output_path, len(rules))
    return output_path


def export_forecast_evidence_to_excel(
    forecast_parquet: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    """Export raw demand forecast rows to a standalone review workbook."""
    forecast_parquet = Path(forecast_parquet)
    if output_path is None:
        output_path = forecast_parquet.parent / "forecast_evidence.xlsx"
    else:
        output_path = Path(output_path)

    table = pq.read_table(str(forecast_parquet))
    rows = table.to_pylist()

    wb = Workbook()
    ws = _active_ws(wb)
    ws.title = "Forecast Evidence"

    preferred_headers = [
        "product_id",
        "facility_id",
        "date",
        "qty_fc",
        "lower_80_qty_fc",
        "upper_80_qty_fc",
    ]
    headers = [name for name in preferred_headers if name in table.column_names]
    if not headers:
        headers = list(table.column_names)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_num, row in enumerate(rows, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(header))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    _auto_size_columns(ws)

    wb.save(str(output_path))
    logger.info("Wrote forecast evidence workbook to %s (%d rows)", output_path, len(rows))
    return output_path


def _write_odoo_import_sheet(wb: Workbook, rules: list[RuleRow]):
    """Write the Odoo-importable sheet.

    Column headers use Odoo's import conventions:
    - `id` for external ID (create/update idempotency)
    - `product_id/.id` for database ID of product
    - `warehouse_id/.id` for database ID of warehouse
    - `location_id/.id` for database ID of stock location
    """
    ws = _active_ws(wb)
    ws.title = "Odoo Import"

    headers = [
        "id",                    # External ID for create/update
        "product_id/.id",        # DB ID of product
        "warehouse_id/.id",      # DB ID of warehouse
        "location_id/.id",       # DB ID of stock location
        "product_min_qty",
        "product_max_qty",
        "trigger",
    ]

    # Write header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT

    # Write data rows (only actionable rules, skip 'skip' actions)
    row_num = 2
    for rule in rules:
        action = rule.get("action")
        if action == "skip":
            continue

        pid = rule.get("_odoo_product_id")
        wid = rule.get("_odoo_warehouse_id")
        lid = rule.get("_odoo_location_id")

        # External ID for idempotent import: mimir.op_{product}_{warehouse}
        external_id = f"mimir.op_{pid}_{wid}"

        ws.cell(row=row_num, column=1, value=external_id)
        ws.cell(row=row_num, column=2, value=pid)
        ws.cell(row=row_num, column=3, value=wid)
        ws.cell(row=row_num, column=4, value=lid)
        ws.cell(row=row_num, column=5, value=rule.get("product_min_qty", 0))
        ws.cell(row=row_num, column=6, value=rule.get("product_max_qty", 0))
        ws.cell(row=row_num, column=7, value=rule.get("trigger", "auto"))
        row_num += 1

    # Auto-size columns
    _auto_size_columns(ws)

    # Add instruction comment at top
    ws.cell(row=row_num + 1, column=1, value="")
    ws.cell(row=row_num + 2, column=1, value="Import instructions:")
    ws.cell(row=row_num + 3, column=1, value="1. In Odoo, go to Inventory > Operations > Replenishment")
    ws.cell(row=row_num + 4, column=1, value='2. Click the gear icon > Import records')
    ws.cell(row=row_num + 5, column=1, value="3. Upload this file and select the 'Odoo Import' sheet")
    ws.cell(row=row_num + 6, column=1, value="4. Verify column mapping, then click Import")
    ws.cell(row=row_num + 7, column=1, value="The 'id' column ensures re-importing updates existing rules.")
    for r in range(row_num + 2, row_num + 8):
        ws.cell(row=r, column=1).font = Font(italic=True, color="666666")


def _write_review_sheet(
    wb: Workbook,
    rules: list[RuleRow],
    decisions_map: dict[tuple[int, int], str],
):
    """Write the human-readable review sheet with all detail."""
    ws = wb.create_sheet("Review")

    headers = [
        ("decision_id", 36),
        ("Product ID", 12),
        ("Product Name", 35),
        ("Action", 10),
        ("product_min_qty", 16),
        ("product_max_qty", 16),
        ("Planned Lead Time (d)", 20),
        ("Empirical Lead Time (d)", 22),
        ("Location ID", 14),
    ]

    # Write header row
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data rows
    for row_num, rule in enumerate(rules, 2):
        pid = rule.get("_odoo_product_id")
        lid = rule.get("_odoo_location_id")
        dec_id = decisions_map.get((pid, lid), "")

        ws.cell(row=row_num, column=1, value=dec_id)
        ws.cell(row=row_num, column=2, value=pid)
        ws.cell(row=row_num, column=3, value=rule.get("product_name", ""))
        ws.cell(row=row_num, column=4, value=rule.get("action", ""))
        ws.cell(row=row_num, column=5, value=rule.get("product_min_qty", 0))
        ws.cell(row=row_num, column=6, value=rule.get("product_max_qty", 0))
        ws.cell(row=row_num, column=7, value=rule.get("planned_lead_time_days", 0))
        ws.cell(row=row_num, column=8, value=rule.get("empirical_lead_time_days", 0))
        ws.cell(row=row_num, column=9, value=lid)

        # Color-code action column
        action = rule.get("action", "")
        action_cell = ws.cell(row=row_num, column=4)
        if action == "skip":
            action_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        elif action == "create":
            action_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif action == "update":
            action_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rules) + 1}"


def _table_to_rows(table: pa.Table, row_type: type[_Row]) -> list[_Row]:
    """Convert a PyArrow table to a list of dicts conforming to `row_type`.

    row_type is a TypedDict declared in schemas.py that mirrors the pa.schema
    of the table being read. pyarrow's to_pylist() returns list[dict[str, Any]]
    which cannot be statically narrowed — the cast asserts the runtime shape
    matches the TypedDict. The contract is enforced by the parquet schema that
    produced the file (see schemas.py); a schema-TypedDict mismatch is a
    class of bug that wouldn't survive the first import anyway.
    """
    return cast(list[_Row], table.to_pylist())


def _auto_size_columns(ws):
    """Auto-size columns based on content width."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
