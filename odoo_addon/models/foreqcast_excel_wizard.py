import base64
import io

from odoo import fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ForeqcastExcelUpload(models.TransientModel):
    _name = "foreqcast.excel.upload"
    _description = "Foreqcast Excel Upload Wizard"

    run_id = fields.Many2one("foreqcast.run", required=True)
    excel_file = fields.Binary("Excel File", required=True)
    filename = fields.Char("Filename")

    def action_upload(self):
        if not openpyxl:
            raise UserError("openpyxl library is required to process Excel files.")

        file_content = base64.b64decode(self.excel_file)
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)

        if "Review" not in wb.sheetnames:
            raise UserError("The uploaded Excel file must contain a sheet named 'Review'.")

        ws = wb["Review"]

        # Read headers
        headers = {cell.value: i for i, cell in enumerate(ws[1])}
        if "decision_id" not in headers:
            raise UserError("The 'Review' sheet must contain a 'decision_id' column.")

        decision_id_idx = headers["decision_id"]
        min_qty_idx = headers.get("product_min_qty")
        max_qty_idx = headers.get("product_max_qty")

        if min_qty_idx is None or max_qty_idx is None:
            raise UserError("The 'Review' sheet must contain 'product_min_qty' and 'product_max_qty' columns.")

        updates = 0
        DecisionRequest = self.env["foreqcast.decision.request"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            decision_id = row[decision_id_idx]
            if not decision_id:
                continue

            try:
                min_qty = float(row[min_qty_idx] or 0)
                max_qty = float(row[max_qty_idx] or 0)
            except (ValueError, TypeError):
                continue

            req = DecisionRequest.search([
                ("run_id", "=", self.run_id.id),
                ("decision_id", "=", decision_id)
            ], limit=1)

            if req and req.status in ("draft", "review"):
                old_min, old_max = req.proposed_min_qty, req.proposed_max_qty

                if old_min != min_qty or old_max != max_qty:
                    req.write({
                        "proposed_min_qty": min_qty,
                        "proposed_max_qty": max_qty,
                        "status": "review"
                    })
                    req.message_post(body=f"Excel Upload updated Min: {old_min} -> {min_qty}, Max: {old_max} -> {max_qty}")
                    updates += 1

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "Excel Upload Complete",
                "message": f"Updated {updates} decision requests from the Excel file.",
                "type": "success",
                "sticky": False,
            },
        }
